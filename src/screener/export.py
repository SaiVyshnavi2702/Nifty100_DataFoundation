import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

from src.screener.composite_score import (
    calculate_sector_relative_score,
    add_historical_fcf_cagr,
)
from src.screener.presets import PRESETS


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "data" / "nifty100.db"
OUTPUT_PATH = PROJECT_ROOT / "output" / "screener_output.xlsx"


KPI_COLUMNS = [
    "company_id",
    "year",
    "broad_sector",
    "market_cap_crore",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "sales",
    "net_profit",
    "return_on_equity_pct",
    "roce_percentage",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "composite_quality_score",
]



GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)


PRESET_THRESHOLDS = {
    "Quality Compounder": {
        "return_on_equity_pct": lambda x: x > 15,
        "debt_to_equity": lambda x: x < 1.0,
        "free_cash_flow_cr": lambda x: x > 0,
        "revenue_cagr_5yr": lambda x: x > 10,
    },

    "Value Pick": {
        "pe_ratio": lambda x: x < 20,
        "pb_ratio": lambda x: x < 3.0,
        "debt_to_equity": lambda x: x < 2.0,
        "dividend_yield_pct": lambda x: x > 1,
    },

    "Growth Accelerator": {
        "pat_cagr_5yr": lambda x: x > 20,
        "revenue_cagr_5yr": lambda x: x > 15,
        "debt_to_equity": lambda x: x < 2.0,
    },

    "Dividend Champion": {
        "dividend_yield_pct": lambda x: x > 2,
        "free_cash_flow_cr": lambda x: x > 0,
    },

    "Debt-Free Blue Chip": {
        "debt_to_equity": lambda x: x == 0,
        "return_on_equity_pct": lambda x: x > 12,
        "sales": lambda x: x > 5000,
    },

    "Turnaround Watch": {
        "free_cash_flow_cr": lambda x: x > 0,
    },
}


def load_export_data():
    """
    Load financial-ratio history together with
    market data, profit/loss data, sector information,
    and ROCE.
    """

    connection = sqlite3.connect(DB_PATH)

    try:
        query = """
            WITH latest_market_cap AS (
                SELECT
                    company_id,
                    market_cap_crore,
                    pe_ratio,
                    pb_ratio,
                    dividend_yield_pct
                FROM (
                    SELECT
                        mc.*,
                        ROW_NUMBER() OVER (
                            PARTITION BY company_id
                            ORDER BY year DESC, id DESC
                        ) AS rn
                    FROM market_cap AS mc
                )
                WHERE rn = 1
            ),

            latest_profit_loss AS (
                SELECT
                    company_id,
                    sales,
                    net_profit
                FROM (
                    SELECT
                        pl.*,
                        ROW_NUMBER() OVER (
                            PARTITION BY company_id
                            ORDER BY year DESC, id DESC
                        ) AS rn
                    FROM profitandloss AS pl
                )
                WHERE rn = 1
            )

            SELECT
                fr.*,
                mc.market_cap_crore,
                mc.pe_ratio,
                mc.pb_ratio,
                mc.dividend_yield_pct,
                pl.sales,
                pl.net_profit,
                s.broad_sector,
                c.roce_percentage
            FROM financial_ratios AS fr
            LEFT JOIN latest_market_cap AS mc
                ON mc.company_id = fr.company_id
            LEFT JOIN latest_profit_loss AS pl
                ON pl.company_id = fr.company_id
            LEFT JOIN sectors AS s
                ON s.company_id = fr.company_id
            LEFT JOIN companies AS c
                ON c.id = fr.company_id
        """

        df = pd.read_sql_query(
            query,
            connection,
        )

    finally:
        connection.close()

    return df


def _latest_per_company(df):
    """
    Keep only the latest financial-ratio row
    for each company.
    """

    result = df.copy()

    result["_year_sort"] = pd.to_numeric(
        result["year"],
        errors="coerce",
    )

    result = (
        result
        .sort_values(
            ["company_id", "_year_sort"],
            ascending=[True, False],
        )
        .drop_duplicates(
            "company_id",
            keep="first",
        )
        .drop(
            columns="_year_sort",
        )
        .reset_index(
            drop=True,
        )
    )

    return result


def prepare_data():
    """
    Prepare latest company-level data.

    Historical FCF CAGR is calculated before
    reducing the data to one row per company.
    """

    df = load_export_data()

    print(
        "Loaded financial-ratio rows:",
        len(df),
    )

    df = add_historical_fcf_cagr(
        df,
        DB_PATH,
    )

    df = _latest_per_company(
        df,
    )

    print(
        "Latest company rows:",
        len(df),
    )

    df = calculate_sector_relative_score(
        df,
    )

    return df


def select_kpis(df):
    """
    Select exactly 20 KPI columns for Excel export.
    """

    if len(KPI_COLUMNS) != 20:
        raise ValueError(
            f"Expected exactly 20 KPI columns, "
            f"but found {len(KPI_COLUMNS)}."
        )

    missing_columns = [
        column
        for column in KPI_COLUMNS
        if column not in df.columns
    ]

    if missing_columns:
        raise KeyError(
            "Missing Day 17 KPI columns: "
            + ", ".join(missing_columns)
        )

    sorted_df = (
        df
        .sort_values(
            by="composite_quality_score",
            ascending=False,
            na_position="last",
        )
        .reset_index(drop=True)
    )

    return sorted_df[KPI_COLUMNS].reset_index(drop=True)


def _cell_passes_threshold(
    preset_name,
    column,
    value,
):
    """
    Check whether a value passes the threshold
    for a particular preset.
    """

    conditions = PRESET_THRESHOLDS.get(
        preset_name,
        {},
    )

    condition = conditions.get(column)

    if condition is None:
        return None

    if pd.isna(value):
        return False

    try:
        return bool(condition(value))
    except (TypeError, ValueError):
        return False


def _apply_excel_formatting(
    worksheet,
    preset_name,
):
    """
    Apply formatting to the Excel worksheet.

    Green cells pass the preset threshold.
    Red cells fail the preset threshold.
    """

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_map = {
        cell.column: cell.value
        for cell in worksheet[1]
    }

    conditions = PRESET_THRESHOLDS.get(
        preset_name,
        {},
    )

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:

            column_name = header_map.get(
                cell.column
            )

            if column_name not in conditions:
                continue

            threshold_result = _cell_passes_threshold(
                preset_name,
                column_name,
                cell.value,
            )

            if threshold_result is True:
                cell.fill = GREEN_FILL

            elif threshold_result is False:
                cell.fill = RED_FILL

    for column_cells in worksheet.columns:

        max_length = 0

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = min(
            max(max_length + 2, 12),
            30,
        )


def export_all_presets():
    """
    Generate the Day 17 Excel workbook.

    Each preset gets its own worksheet.
    Every worksheet contains exactly 20 columns.
    """

    OUTPUT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    df = prepare_data()

    print()
    print(
        "Day 17 export preparation complete."
    )
    print(
        "Companies scored:",
        len(df),
    )

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:

        for preset_name, preset_function in PRESETS.items():

            print(
                f"Processing preset: {preset_name}"
            )

            preset_result = preset_function(
                df,
            )

            preset_result = select_kpis(
                preset_result,
            )

            sheet_name = preset_name[:31]

            preset_result.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

            worksheet = writer.sheets[
                sheet_name
            ]

            _apply_excel_formatting(
                worksheet,
                preset_name,
            )

            print(
                f"{preset_name}: "
                f"{len(preset_result)} companies"
            )

    print()
    print(
        "Day 17 export completed successfully."
    )
    print(
        "Output:",
        OUTPUT_PATH,
    )


if __name__ == "__main__":
    export_all_presets()
