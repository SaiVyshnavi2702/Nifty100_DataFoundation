from pathlib import Path
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]

DB_PATH = PROJECT_ROOT / "data" / "nifty100.db"

PEER_GROUPS_FILE = (
    PROJECT_ROOT
    / "data"
    / "raw"
    / "supporting"
    / "peer_groups.xlsx"
)

OUTPUT_DIR = PROJECT_ROOT / "output"

OUTPUT_FILE = OUTPUT_DIR / "peer_comparison.xlsx"


METRICS = [
    "ROE",
    "ROCE",
    "Net Profit Margin",
    "D/E",
    "FCF",
    "PAT CAGR 5yr",
    "Revenue CAGR 5yr",
    "EPS CAGR 5yr",
    "Interest Coverage",
    "Asset Turnover",
]


METRIC_COLUMNS = {
    "ROE": "return_on_equity_pct",
    "ROCE": "roce_percentage",
    "Net Profit Margin": "net_profit_margin_pct",
    "D/E": "debt_to_equity",
    "FCF": "free_cash_flow_cr",
    "PAT CAGR 5yr": "pat_cagr_5yr",
    "Revenue CAGR 5yr": "revenue_cagr_5yr",
    "EPS CAGR 5yr": "eps_cagr_5yr",
    "Interest Coverage": "interest_coverage",
    "Asset Turnover": "asset_turnover",
}


PERCENTILE_COLUMNS = {
    metric: f"{metric} Percentile"
    for metric in METRICS
}


GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

SUMMARY_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)


def load_peer_groups():
    if not PEER_GROUPS_FILE.exists():
        raise FileNotFoundError(
            f"Peer group file not found: {PEER_GROUPS_FILE}"
        )

    df = pd.read_excel(
        PEER_GROUPS_FILE
    )

    required_columns = {
        "peer_group_name",
        "company_id",
        "is_benchmark",
    }

    missing = required_columns - set(df.columns)

    if missing:
        raise ValueError(
            "Missing columns in peer_groups.xlsx: "
            + ", ".join(sorted(missing))
        )

    df = df[
        [
            "peer_group_name",
            "company_id",
            "is_benchmark",
        ]
    ].copy()

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
    )

    df["peer_group_name"] = (
        df["peer_group_name"]
        .astype(str)
        .str.strip()
    )

    df["is_benchmark"] = (
        df["is_benchmark"]
        .astype(str)
        .str.lower()
        .isin(
            [
                "true",
                "1",
                "yes",
            ]
        )
    )

    df = df[
        (df["company_id"] != "")
        & (df["peer_group_name"] != "")
    ].copy()

    benchmark_counts = (
        df.groupby(
            "peer_group_name"
        )["is_benchmark"]
        .sum()
    )

    invalid_groups = benchmark_counts[
        benchmark_counts != 1
    ]

    if not invalid_groups.empty:
        raise ValueError(
            "Each peer group must have exactly one benchmark company. "
            f"Invalid groups: {invalid_groups.index.tolist()}"
        )

    return df


def load_companies(conn):
    query = """
        SELECT
            id AS company_id,
            company_name
        FROM companies
    """

    df = pd.read_sql_query(
        query,
        conn,
    )

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
    )

    return df


def load_financial_data(conn):
    query = """
        SELECT
            fr.company_id,
            fr.year,
            fr.return_on_equity_pct,
            fr.net_profit_margin_pct,
            fr.debt_to_equity,
            fr.free_cash_flow_cr,
            fr.pat_cagr_5yr,
            fr.revenue_cagr_5yr,
            fr.eps_cagr_5yr,
            fr.interest_coverage,
            fr.asset_turnover,

            pl.operating_profit,
            pl.other_income,

            bs.equity_capital,
            bs.reserves,
            bs.borrowings

        FROM financial_ratios AS fr

        LEFT JOIN profitandloss AS pl
            ON pl.company_id = fr.company_id
            AND pl.year = fr.year
            AND pl.period = fr.period

        LEFT JOIN balancesheet AS bs
            ON bs.company_id = fr.company_id
            AND bs.year = fr.year
            AND bs.period = fr.period
    """

    df = pd.read_sql_query(
        query,
        conn,
    )

    if df.empty:
        raise ValueError(
            "No financial data found."
        )

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
    )

    df["year"] = pd.to_numeric(
        df["year"],
        errors="coerce",
    )

    numeric_columns = [
        "return_on_equity_pct",
        "net_profit_margin_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "pat_cagr_5yr",
        "revenue_cagr_5yr",
        "eps_cagr_5yr",
        "interest_coverage",
        "asset_turnover",
        "operating_profit",
        "other_income",
        "equity_capital",
        "reserves",
        "borrowings",
    ]

    for column in numeric_columns:
        df[column] = pd.to_numeric(
            df[column],
            errors="coerce",
        )

    ebit = (
        df["operating_profit"]
        + df["other_income"]
    )

    capital_employed = (
        df["equity_capital"]
        + df["reserves"]
        + df["borrowings"]
    )

    df["roce_percentage"] = (
        ebit
        / capital_employed
        * 100
    )

    df["roce_percentage"] = (
        df["roce_percentage"]
        .replace(
            [float("inf"), float("-inf")],
            pd.NA,
        )
    )

    return df


def select_latest_data(financial_data):
    data = financial_data.copy()

    data = data.dropna(
        subset=["company_id", "year"]
    )

    data = data.sort_values(
        [
            "company_id",
            "year",
        ]
    )

    latest = (
        data.groupby(
            "company_id",
            as_index=False,
        )
        .tail(1)
        .copy()
    )

    return latest


def load_day18_percentiles(conn):
    query = """
        SELECT
            company_id,
            peer_group_name,
            metric,
            percentile_rank,
            year
        FROM peer_percentiles
    """

    df = pd.read_sql_query(
        query,
        conn,
    )

    if df.empty:
        raise ValueError(
            "peer_percentiles table is empty. "
            "Run Day 18 first."
        )

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
    )

    df["year"] = pd.to_numeric(
        df["year"],
        errors="coerce",
    )

    df["percentile_rank"] = pd.to_numeric(
        df["percentile_rank"],
        errors="coerce",
    )

    return df


def prepare_data(
    peer_groups,
    companies,
    financial_data,
    percentile_data,
):
    latest_financial = select_latest_data(
        financial_data
    )

    latest_year = (
        latest_financial[
            [
                "company_id",
                "year",
            ]
        ]
        .copy()
    )

    latest_year = latest_year.rename(
        columns={
            "year": "latest_year"
        }
    )

    company_data = peer_groups.merge(
        companies,
        on="company_id",
        how="left",
    )

    company_data = company_data.merge(
        latest_year,
        on="company_id",
        how="left",
    )

    company_data = company_data.merge(
        latest_financial[
            [
                "company_id",
                "year",
                *METRIC_COLUMNS.values(),
            ]
        ],
        left_on=[
            "company_id",
            "latest_year",
        ],
        right_on=[
            "company_id",
            "year",
        ],
        how="left",
    )

    company_data = company_data.drop(
        columns=["year"],
        errors="ignore",
    )

    percentile_data = percentile_data.copy()

    percentile_data = percentile_data.merge(
        peer_groups[
            [
                "company_id",
                "peer_group_name",
            ]
        ],
        on=[
            "company_id",
            "peer_group_name",
        ],
        how="inner",
    )

    percentile_data = percentile_data.merge(
        latest_year,
        on="company_id",
        how="left",
    )

    percentile_data = percentile_data[
        percentile_data["year"]
        == percentile_data["latest_year"]
    ].copy()

    return company_data, percentile_data


def build_peer_sheet(
    peer_group_name,
    company_data,
    percentile_data,
):
    group_companies = company_data[
        company_data["peer_group_name"]
        == peer_group_name
    ].copy()

    if group_companies.empty:
        return pd.DataFrame()

    rows = []

    for _, company in group_companies.iterrows():

        row = {
            "company_id": company["company_id"],
            "company_name": company["company_name"],
        }

        for metric, column in METRIC_COLUMNS.items():
            row[metric] = company[column]

        company_id = company["company_id"]
        year = company["latest_year"]

        company_percentiles = percentile_data[
            (percentile_data["company_id"] == company_id)
            & (percentile_data["year"] == year)
        ]

        for metric in METRICS:

            match = company_percentiles[
                company_percentiles["metric"] == metric
            ]

            if match.empty:
                row[
                    PERCENTILE_COLUMNS[metric]
                ] = pd.NA
            else:
                row[
                    PERCENTILE_COLUMNS[metric]
                ] = match[
                    "percentile_rank"
                ].iloc[0]

        rows.append(row)

    columns = [
        "company_id",
        "company_name",
        *METRICS,
        *PERCENTILE_COLUMNS.values(),
    ]

    return pd.DataFrame(
        rows,
        columns=columns,
    )


def add_summary_row(
    df,
):
    summary = {}

    summary["company_id"] = "PEER GROUP MEDIAN"
    summary["company_name"] = ""

    for metric in METRICS:
        summary[metric] = pd.to_numeric(
            df[metric],
            errors="coerce",
        ).median()

    for column in PERCENTILE_COLUMNS.values():
        summary[column] = pd.to_numeric(
            df[column],
            errors="coerce",
        ).median()

    return pd.DataFrame(
        [summary],
        columns=df.columns,
    )


def safe_sheet_name(name):
    invalid_characters = [
        "\\",
        "/",
        "*",
        "?",
        ":",
        "[",
        "]",
    ]

    result = str(name)

    for character in invalid_characters:
        result = result.replace(
            character,
            "_",
        )

    return result[:31]


def write_excel(
    peer_groups,
    company_data,
    percentile_data,
):
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    if OUTPUT_FILE.exists():
        OUTPUT_FILE.unlink()

    group_names = (
        peer_groups["peer_group_name"]
        .drop_duplicates()
        .tolist()
    )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        for peer_group_name in group_names:

            sheet_data = build_peer_sheet(
                peer_group_name,
                company_data,
                percentile_data,
            )

            if sheet_data.empty:
                continue

            summary = add_summary_row(
                sheet_data
            )

            final_data = pd.concat(
                [
                    sheet_data,
                    summary,
                ],
                ignore_index=True,
            )

            sheet_name = safe_sheet_name(
                peer_group_name
            )

            final_data.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

    format_excel(
        peer_groups,
    )


def format_excel(peer_groups):
    workbook = load_workbook(
        OUTPUT_FILE
    )

    benchmark_lookup = dict(
        zip(
            peer_groups["company_id"],
            peer_groups["is_benchmark"],
        )
    )

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "C2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        percentile_columns = []

        for metric in METRICS:
            column_name = PERCENTILE_COLUMNS[metric]

            if column_name in headers:
                percentile_columns.append(
                    headers[column_name]
                )

        summary_row = worksheet.max_row

        for row in range(
            2,
            summary_row,
        ):

            company_id = worksheet.cell(
                row=row,
                column=1,
            ).value

            is_benchmark = benchmark_lookup.get(
                str(company_id),
                False,
            )

            if is_benchmark:
                for column in range(
                    1,
                    worksheet.max_column + 1,
                ):
                    worksheet.cell(
                        row=row,
                        column=column,
                    ).fill = BENCHMARK_FILL

            for column in percentile_columns:

                cell = worksheet.cell(
                    row=row,
                    column=column,
                )

                value = cell.value

                if value is None:
                    continue

                try:
                    value = float(value)
                except (
                    TypeError,
                    ValueError,
                ):
                    continue

                if value >= 0.75:
                    cell.fill = GREEN_FILL

                elif value > 0.25:
                    cell.fill = YELLOW_FILL

                else:
                    cell.fill = RED_FILL

        for column in range(
            1,
            worksheet.max_column + 1,
        ):
            cell = worksheet.cell(
                row=summary_row,
                column=column,
            )

            cell.fill = SUMMARY_FILL
            cell.font = Font(
                bold=True
            )

        for column in percentile_columns:

            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = 18

        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 30

        for column in range(
            3,
            worksheet.max_column + 1,
        ):
            if column not in percentile_columns:
                worksheet.column_dimensions[
                    get_column_letter(column)
                ].width = 18

        for row in range(
            2,
            worksheet.max_row + 1,
        ):
            for column in percentile_columns:
                cell = worksheet.cell(
                    row=row,
                    column=column,
                )

                if (
                    cell.value is not None
                    and isinstance(
                        cell.value,
                        (int, float),
                    )
                ):
                    cell.number_format = "0.0%"

        worksheet.row_dimensions[1].height = 30

    workbook.save(
        OUTPUT_FILE
    )


def validate_output(
    peer_groups,
):
    if not OUTPUT_FILE.exists():
        raise AssertionError(
            "peer_comparison.xlsx was not created."
        )

    workbook = load_workbook(
        OUTPUT_FILE,
        read_only=True,
    )

    expected_groups = set(
        peer_groups["peer_group_name"]
    )

    actual_sheets = set(
        workbook.sheetnames
    )

    expected_sheets = {
        safe_sheet_name(name)
        for name in expected_groups
    }

    if actual_sheets != expected_sheets:
        raise AssertionError(
            "Peer group sheets do not match.\n"
            f"Expected: {sorted(expected_sheets)}\n"
            f"Actual: {sorted(actual_sheets)}"
        )

    for sheet_name in workbook.sheetnames:

        worksheet = workbook[sheet_name]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        expected_columns = [
            "company_id",
            "company_name",
            *METRICS,
            *PERCENTILE_COLUMNS.values(),
        ]

        if headers != expected_columns:
            raise AssertionError(
                f"Incorrect columns in sheet: {sheet_name}"
            )

        if worksheet.max_row < 2:
            raise AssertionError(
                f"Sheet is empty: {sheet_name}"
            )

    workbook.close()

    print(
        f"Excel file created: {OUTPUT_FILE}"
    )

    print(
        f"Sheets created: {len(actual_sheets)}"
    )


def main():
    print("Day 20 - Peer Comparison Excel Report")
    print("--------------------------------------")

    print()
    print("Loading peer groups...")

    peer_groups = load_peer_groups()

    print(
        f"Peer groups loaded: "
        f"{peer_groups['peer_group_name'].nunique()}"
    )

    print()
    print("Opening database...")

    conn = sqlite3.connect(
        DB_PATH
    )

    try:
        print()
        print("Loading companies...")

        companies = load_companies(
            conn
        )

        print(
            f"Companies loaded: "
            f"{len(companies)}"
        )

        print()
        print("Loading financial data...")

        financial_data = load_financial_data(
            conn
        )

        print(
            f"Financial rows loaded: "
            f"{len(financial_data)}"
        )

        print()
        print("Loading Day 18 percentile data...")

        percentile_data = load_day18_percentiles(
            conn
        )

        print(
            f"Percentile rows loaded: "
            f"{len(percentile_data)}"
        )

        print()
        print("Preparing comparison data...")

        company_data, percentile_data = prepare_data(
            peer_groups,
            companies,
            financial_data,
            percentile_data,
        )

        print(
            f"Companies prepared: "
            f"{len(company_data)}"
        )

        print()
        print("Creating Excel report...")

        write_excel(
            peer_groups,
            company_data,
            percentile_data,
        )

        print()
        print("Validating Excel report...")

        validate_output(
            peer_groups
        )

        print()
        print("Day 20 completed successfully.")

        print(
            f"Output file: "
            f"{OUTPUT_FILE}"
        )

    finally:
        conn.close()


if __name__ == "__main__":
    main()
